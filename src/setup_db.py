import sqlite3
import openpyxl

# Open the Excel workbook
workbook = openpyxl.load_workbook('src/db.xlsx')

# Connect to the SQLite database
conn = sqlite3.connect('instance/db.sqlite3')
cursor = conn.cursor()

# Loop through each sheet in the workbook
for sheet in workbook.sheetnames:
    # Get the worksheet and the column headers
    worksheet = workbook[sheet]
    headers = [cell.value for cell in worksheet[1] if cell.value is not None]

    # Build the SQL query to create the table
    columns = ", ".join(headers)
    query = f"CREATE TABLE IF NOT EXISTS {sheet} ({columns})"

    # Create the table in the database
    cursor.execute(query)
    conn.commit()

    # Loop through each row in the worksheet and insert the data into the table
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Filter out None values from row
        row = [cell_value for cell_value in row if cell_value is not None]
        query = f"INSERT INTO {sheet} ({columns}) VALUES ({','.join(['?' for _ in row])})"
        cursor.execute(query, row)
        conn.commit()

# Close the database connection
cursor.close()
conn.close()
